from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from datetime import datetime
import openpyxl
import time

options = Options()
# options.add_argument("start-maximized")
# options.add_experimental_option("detach", True)
options.add_argument("--lang=en-GB")
options.add_experimental_option('excludeSwitches', ['enable-logging'])
driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)

weekday = datetime.now().strftime('%A')
workbook = openpyxl.load_workbook('Data.xlsx')
sheets = workbook.sheetnames

for sheet in sheets:
    if weekday == sheet:
        x = workbook[sheet]
        # y = list(x.values)
        break

for i in range (1, 13):
    c = x.cell(row = i, column = 3)
    if c.value is not None:
        search = c.value

        driver.get("https://www.google.com")
        driver.find_element(By.NAME, 'q').send_keys(search)
        time.sleep(2)

        content = driver.find_elements(By.XPATH, "//form[@action='/search' and @role='search']//ul[@role='listbox']//li")

        auto_suggetion = dict()

        for element in content:
            temp = element.text
            if "\n" in temp:
                auto_suggetion[temp] = int(len(temp) - 2)
            else:
                auto_suggetion[temp] = int(len(temp))
        
        max_value = max(auto_suggetion, key=auto_suggetion.get)
        min_value = min(auto_suggetion, key=auto_suggetion.get)

        c1 = x.cell(row = i, column = 4)
        c1.value = max_value

        c2 = x.cell(row = i, column = 5)
        c2.value = min_value

workbook.save('Data.xlsx')


